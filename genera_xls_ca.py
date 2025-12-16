# genera_xls_ca.py
# Dado un código de compra ágil, genera un archivo Excel

import os
import pandas as pd
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import descarga_ca


def _cargar_manifest_adjuntos(codigo_ca):
    carpeta_base = os.path.join("Descargas", "ComprasAgiles", codigo_ca)
    ruta_manifest = os.path.join(carpeta_base, descarga_ca.MANIFEST_ADJUNTOS_FILENAME)
    if not os.path.exists(ruta_manifest):
        return {}
    try:
        with open(ruta_manifest, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return {}

    proveedores = data.get("proveedores") or []
    by_rut = {}
    for p in proveedores:
        rut = p.get("rut")
        rut_norm = descarga_ca._normalizar_rut(rut) or (str(rut).strip() if rut else None)
        if not rut_norm:
            continue
        by_rut[rut_norm] = p
    return by_rut

def generar_excel_compra_agil(codigo_ca, driver=None):
    """
    Genera un archivo Excel con la información de una compra ágil
    
    Args:
        codigo_ca (str): Código de la compra ágil
        driver: Instancia del navegador Selenium
    
    Returns:
        str: Ruta del archivo Excel generado
    """
    if not driver:
        print("Error: Se requiere una instancia del navegador")
        return None
    
    try:
        print(f"Generando Excel para compra ágil: {codigo_ca}")
        
        # Navegar a la compra ágil
        if not descarga_ca.navegar_a_compra_agil(codigo_ca, driver):
            print("Error al navegar a la compra ágil")
            return None
        
        # Extraer información general de la compra ágil
        info_compra = extraer_informacion_compra_agil(codigo_ca, driver)
        
        # Extraer datos de proveedores
        datos_proveedores = extraer_datos_proveedores(codigo_ca, driver)
        
        if not datos_proveedores:
            print("No se encontraron proveedores para generar el Excel")
            return None
        
        # Crear el archivo Excel
        ruta_excel = crear_estructura_excel(datos_proveedores, codigo_ca, info_compra)
        
        if ruta_excel:
            print(f"Excel generado exitosamente: {ruta_excel}")
            return ruta_excel
        else:
            print("Error al crear el archivo Excel")
            return None
            
    except Exception as e:
        print(f"Error al generar Excel: {str(e)}")
        return None

def extraer_informacion_compra_agil(codigo_ca, driver):
    """
    Extrae información general de la compra ágil
    
    Args:
        codigo_ca (str): Código de la compra ágil
        driver: Instancia del navegador Selenium
    
    Returns:
        dict: Diccionario con información de la compra ágil
    """
    info_compra = {
        'codigo': codigo_ca,
        'nombre': '',
        'organismo': '',
        'fecha_publicacion': '',
        'fecha_cierre': '',
        'estado': '',
        'monto_estimado': '',
        'descripcion': ''
    }
    
    try:
        # Extraer título/nombre de la compra ágil
        try:
            elemento_titulo = driver.find_element(By.XPATH, "//h1 | //h2 | //div[contains(@class, 'title')] | //span[contains(@class, 'title')]")
            info_compra['nombre'] = elemento_titulo.text.strip()
        except:
            info_compra['nombre'] = f"Compra Ágil {codigo_ca}"
        
        # Extraer organismo
        try:
            elemento_organismo = driver.find_element(By.XPATH, "//td[contains(text(), 'Organismo')] | //span[contains(text(), 'Organismo')]")
            info_compra['organismo'] = elemento_organismo.text.replace('Organismo:', '').strip()
        except:
            pass
        
        # Extraer fechas
        try:
            elemento_fecha_pub = driver.find_element(By.XPATH, "//td[contains(text(), 'Publicación')] | //span[contains(text(), 'Publicación')]")
            info_compra['fecha_publicacion'] = elemento_fecha_pub.text.replace('Publicación:', '').strip()
        except:
            pass
        
        try:
            elemento_fecha_cierre = driver.find_element(By.XPATH, "//td[contains(text(), 'Cierre')] | //span[contains(text(), 'Cierre')]")
            info_compra['fecha_cierre'] = elemento_fecha_cierre.text.replace('Cierre:', '').strip()
        except:
            pass
        
        # Extraer estado
        try:
            elemento_estado = driver.find_element(By.XPATH, "//td[contains(text(), 'Estado')] | //span[contains(text(), 'Estado')]")
            info_compra['estado'] = elemento_estado.text.replace('Estado:', '').strip()
        except:
            pass
        
        # Extraer monto estimado
        try:
            elemento_monto = driver.find_element(By.XPATH, "//td[contains(text(), 'Monto')] | //span[contains(text(), 'Monto')]")
            info_compra['monto_estimado'] = elemento_monto.text.replace('Monto:', '').strip()
        except:
            pass
        
        return info_compra
        
    except Exception as e:
        print(f"Error al extraer información de la compra ágil: {str(e)}")
        return info_compra

def extraer_datos_proveedores(codigo_ca, driver):
    """
    Extrae los datos de los proveedores participantes
    
    Args:
        codigo_ca (str): Código de la compra ágil
        driver: Instancia del navegador Selenium
    
    Returns:
        list: Lista de diccionarios con datos de proveedores
    """
    try:
        # Obtener proveedores usando la función del módulo de descarga
        proveedores = descarga_ca.obtener_proveedores_ca(driver)
        manifest_by_rut = _cargar_manifest_adjuntos(codigo_ca)
        
        datos_proveedores = []
        
        for i, proveedor in enumerate(proveedores, 1):
            # Construir rutas basadas en la estructura de carpetas
            carpeta_base = os.path.join("Descargas", "ComprasAgiles", codigo_ca)
            rut_norm = descarga_ca._normalizar_rut(proveedor.get("rut")) or (proveedor.get("rut") or "").strip()
            entry = manifest_by_rut.get(rut_norm) if rut_norm else None

            carpeta_proveedor = None
            if entry and entry.get("carpeta"):
                carpeta_proveedor = entry.get("carpeta")
            else:
                nombre_proveedor_limpio = descarga_ca.limpiar_nombre_archivo(proveedor['nombre'])
                carpeta_proveedor = os.path.join(carpeta_base, nombre_proveedor_limpio)

            nombre_carpeta = os.path.basename(carpeta_proveedor.rstrip(os.sep)) if carpeta_proveedor else ""
            ruta_zip = os.path.join(carpeta_base, f"{nombre_carpeta}.zip") if nombre_carpeta else "No creado"

            adjuntos_esperados = None
            if entry:
                adjuntos_esperados = entry.get("esperados_ui")
                if adjuntos_esperados is None:
                    adjuntos_esperados = entry.get("esperados_api")
            
            # Verificar si existen los archivos/carpetas
            existe_carpeta = os.path.exists(carpeta_proveedor)
            existe_zip = os.path.exists(ruta_zip)
            
            # Contar adjuntos si existe la carpeta
            num_adjuntos = 0
            if existe_carpeta:
                try:
                    archivos = [f for f in os.listdir(carpeta_proveedor) if os.path.isfile(os.path.join(carpeta_proveedor, f))]
                    num_adjuntos = len(archivos)
                except:
                    pass
            
            datos_proveedor = {
                'N°': i,
                'Nombre Proveedor': proveedor['nombre'],
                'RUT': proveedor['rut'],
                'Carpeta Path': carpeta_proveedor if existe_carpeta else 'No descargada',
                'ZIP Path': ruta_zip if existe_zip else 'No creado',
                'Adjuntos Esperados': adjuntos_esperados if adjuntos_esperados is not None else '',
                'Número Adjuntos': num_adjuntos,
                'Estado Descarga': 'Completada' if existe_carpeta and num_adjuntos > 0 else 'Pendiente',
                'Estado ZIP': 'Creado' if existe_zip else 'No creado',
                'Fecha Procesamiento': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            datos_proveedores.append(datos_proveedor)
        
        return datos_proveedores
        
    except Exception as e:
        print(f"Error al extraer datos de proveedores: {str(e)}")
        return []

def crear_estructura_excel(datos_proveedores, codigo_ca, info_compra):
    """
    Crea la estructura del archivo Excel
    
    Args:
        datos_proveedores (list): Lista con datos de proveedores
        codigo_ca (str): Código de la compra ágil
        info_compra (dict): Información general de la compra ágil
    
    Returns:
        str: Ruta del archivo Excel creado
    """
    try:
        # Crear carpeta de destino
        carpeta_destino = os.path.join("Descargas", "ComprasAgiles", codigo_ca)
        os.makedirs(carpeta_destino, exist_ok=True)
        
        # Nombre del archivo Excel
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_excel = f"CompraAgil_{codigo_ca}_{timestamp}.xlsx"
        ruta_excel = os.path.join(carpeta_destino, nombre_excel)
        
        # Crear workbook
        wb = Workbook()
        
        # Crear hoja de información general
        ws_info = wb.active
        ws_info.title = "Información General"
        crear_hoja_informacion(ws_info, info_compra, codigo_ca)
        
        # Crear hoja de proveedores
        ws_proveedores = wb.create_sheet("Proveedores")
        crear_hoja_proveedores(ws_proveedores, datos_proveedores)
        
        # Crear hoja de resumen
        ws_resumen = wb.create_sheet("Resumen")
        crear_hoja_resumen(ws_resumen, datos_proveedores, info_compra)
        
        # Guardar archivo
        wb.save(ruta_excel)
        
        print(f"Excel creado: {nombre_excel}")
        return ruta_excel
        
    except Exception as e:
        print(f"Error al crear estructura Excel: {str(e)}")
        return None

def crear_hoja_informacion(ws, info_compra, codigo_ca):
    """
    Crea la hoja de información general
    """
    # Título
    ws['A1'] = f"COMPRA ÁGIL - {codigo_ca}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws.merge_cells('A1:B1')
    
    # Información
    row = 3
    campos = [
        ('Código:', info_compra['codigo']),
        ('Nombre:', info_compra['nombre']),
        ('Organismo:', info_compra['organismo']),
        ('Fecha Publicación:', info_compra['fecha_publicacion']),
        ('Fecha Cierre:', info_compra['fecha_cierre']),
        ('Estado:', info_compra['estado']),
        ('Monto Estimado:', info_compra['monto_estimado']),
        ('Fecha Generación Excel:', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    ]
    
    for campo, valor in campos:
        ws[f'A{row}'] = campo
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = valor
        row += 1
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50

def crear_hoja_proveedores(ws, datos_proveedores):
    """
    Crea la hoja de proveedores
    """
    # Título
    ws['A1'] = "LISTA DE PROVEEDORES"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws.merge_cells('A1:J1')
    
    # Encabezados
    encabezados = ['N°', 'Nombre Proveedor', 'RUT', 'Carpeta Path', 'ZIP Path', 
                   'Adjuntos Esperados', 'Número Adjuntos', 'Estado Descarga', 'Estado ZIP', 'Fecha Procesamiento']
    
    for col, encabezado in enumerate(encabezados, 1):
        cell = ws.cell(row=3, column=col, value=encabezado)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Datos
    for row_idx, proveedor in enumerate(datos_proveedores, 4):
        ws.cell(row=row_idx, column=1, value=proveedor['N°'])
        ws.cell(row=row_idx, column=2, value=proveedor['Nombre Proveedor'])
        ws.cell(row=row_idx, column=3, value=proveedor['RUT'])
        ws.cell(row=row_idx, column=4, value=proveedor['Carpeta Path'])
        ws.cell(row=row_idx, column=5, value=proveedor['ZIP Path'])
        ws.cell(row=row_idx, column=6, value=proveedor.get('Adjuntos Esperados', ''))
        ws.cell(row=row_idx, column=7, value=proveedor['Número Adjuntos'])
        ws.cell(row=row_idx, column=8, value=proveedor['Estado Descarga'])
        ws.cell(row=row_idx, column=9, value=proveedor['Estado ZIP'])
        ws.cell(row=row_idx, column=10, value=proveedor['Fecha Procesamiento'])
    
    # Ajustar anchos de columna
    anchos = [5, 30, 15, 40, 40, 18, 15, 15, 15, 20]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[chr(64 + col)].width = ancho
    
    # Aplicar bordes
    aplicar_bordes_tabla(ws, 3, len(datos_proveedores) + 3, len(encabezados))

def crear_hoja_resumen(ws, datos_proveedores, info_compra):
    """
    Crea la hoja de resumen
    """
    # Título
    ws['A1'] = "RESUMEN DE PROCESAMIENTO"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws.merge_cells('A1:B1')
    
    # Estadísticas
    total_proveedores = len(datos_proveedores)
    proveedores_descargados = sum(1 for p in datos_proveedores if p['Estado Descarga'] == 'Completada')
    zips_creados = sum(1 for p in datos_proveedores if p['Estado ZIP'] == 'Creado')
    total_adjuntos = sum(p['Número Adjuntos'] for p in datos_proveedores)
    total_adjuntos_esperados = 0
    for p in datos_proveedores:
        v = p.get('Adjuntos Esperados')
        if isinstance(v, int):
            total_adjuntos_esperados += v
        else:
            try:
                if str(v).strip():
                    total_adjuntos_esperados += int(v)
            except Exception:
                pass
    
    row = 3
    estadisticas = [
        ('Total Proveedores:', total_proveedores),
        ('Proveedores Descargados:', proveedores_descargados),
        ('ZIPs Creados:', zips_creados),
        ('Total Adjuntos Esperados:', total_adjuntos_esperados),
        ('Total Adjuntos:', total_adjuntos),
        ('Porcentaje Completado:', f"{(proveedores_descargados/total_proveedores*100):.1f}%" if total_proveedores > 0 else "0%")
    ]
    
    for estadistica, valor in estadisticas:
        ws[f'A{row}'] = estadistica
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = valor
        row += 1
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20

def aplicar_bordes_tabla(ws, fila_inicio, fila_fin, num_columnas):
    """
    Aplica bordes a una tabla
    """
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(fila_inicio, fila_fin + 1):
        for col in range(1, num_columnas + 1):
            ws.cell(row=row, column=col).border = thin_border

if __name__ == "__main__":
    # Función principal para pruebas
    print("Módulo genera_xls_ca.py - Generación de Excel para compras ágiles")
    print("Este módulo debe ser importado y usado desde app.py")
