import json
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def generar_excel_licitacion(codigo_lici, driver=None, resumen=None, manifest_path=None):
    """
    Genera un Excel sencillo de resumen para una licitacion.

    Prioriza un resumen ya calculado (por ejemplo, retornado por flujo_licitacion.test_flujo_licitacion).
    Si no se entrega, intenta leer un manifest_licitacion.json en la carpeta de Descargas.
    """
    data = resumen or _cargar_manifest(codigo_lici, manifest_path)
    proveedores = data.get("proveedores") if isinstance(data, dict) else None
    if not proveedores:
        return None

    carpeta = os.path.join("Descargas", "Licitaciones", codigo_lici)
    os.makedirs(carpeta, exist_ok=True)
    ruta_excel = os.path.join(carpeta, f"resumen_{codigo_lici}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    _render_encabezado(ws, codigo_lici)

    headers = ["Proveedor", "RUT", "Total adjuntos", "Administrativos", "Tecnicos", "Economicos", "Carpeta"]
    ws.append(headers)
    _pintar_encabezado(ws, ws.max_row, len(headers))

    for prov in proveedores:
        counts = _contar_adjuntos(prov)
        ws.append(
            [
                prov.get("nombre") or "",
                prov.get("rut") or "",
                counts["total"],
                counts["admin"],
                counts["tecnico"],
                counts["economico"],
                prov.get("carpeta") or "",
            ]
        )

    _ajustar_columnas(ws, [28, 16, 16, 16, 16, 16, 48])

    errores = data.get("errores") or []
    _render_errores_sheet(wb, proveedores, errores)

    wb.save(ruta_excel)
    return ruta_excel


# ---------------- internal helpers ----------------
def _cargar_manifest(codigo_lici, manifest_path=None):
    ruta = manifest_path or os.path.join("Descargas", "Licitaciones", codigo_lici, "manifest_licitacion.json")
    if not os.path.exists(ruta):
        return {}
    try:
        with open(ruta, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _contar_adjuntos(prov_entry):
    def _contar_carpeta(carpeta_tipo):
        total = 0
        if os.path.isdir(carpeta_tipo):
            for _, _, files in os.walk(carpeta_tipo):
                total += len(files)
        return total

    counts = {
        "admin": (prov_entry.get("admin") or {}).get("descargados") or 0,
        "tecnico": (prov_entry.get("tecnico") or {}).get("descargados") or 0,
        "economico": (prov_entry.get("economico") or {}).get("descargados") or 0,
    }

    carpeta_base = prov_entry.get("carpeta")
    if carpeta_base:
        if counts["admin"] == 0:
            counts["admin"] = _contar_carpeta(os.path.join(carpeta_base, "ADMINISTRATIVOS"))
        if counts["tecnico"] == 0:
            counts["tecnico"] = _contar_carpeta(os.path.join(carpeta_base, "TECNICOS"))
        if counts["economico"] == 0:
            counts["economico"] = _contar_carpeta(os.path.join(carpeta_base, "ECONOMICOS"))

    total = prov_entry.get("total_descargados")
    if total is None:
        total = counts["admin"] + counts["tecnico"] + counts["economico"]
    return {"total": total, **counts}


def _render_encabezado(ws, codigo):
    ws.append(["Licitacion", codigo])
    ws.append(["Generado", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append([])  # linea en blanco


def _pintar_encabezado(ws, row_index, num_cols):
    fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    bold = Font(bold=True)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_index, column=col)
        cell.fill = fill
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _ajustar_columnas(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width


def _render_errores_sheet(wb, proveedores, errores_generales):
    ws_err = wb.create_sheet("Errores")
    ws_err.append(["Origen", "Detalle"])
    _pintar_encabezado(ws_err, 1, 2)

    for err in errores_generales:
        ws_err.append(["General", err])

    for prov in proveedores:
        nombre = prov.get("nombre") or prov.get("rut") or "Proveedor"
        for etiqueta, info in [
            ("Administrativos", prov.get("admin") or {}),
            ("Tecnicos", prov.get("tecnico") or {}),
            ("Economicos", prov.get("economico") or {}),
        ]:
            for err in info.get("errores") or []:
                ws_err.append([nombre, f"{etiqueta}: {err}"])
